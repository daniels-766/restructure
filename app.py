import logging
import gspread
from datetime import datetime
import locale
from google.oauth2.service_account import Credentials
from decimal import Decimal
from sqlalchemy import func, extract
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, g, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
from datetime import datetime, timedelta, date
import pytz
import re
from sqlalchemy import or_, extract, func
import calendar
from werkzeug.utils import secure_filename
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
import uuid
from io import BytesIO
from math import ceil
from flask import send_file, request
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from sqlalchemy.dialects.mysql import LONGTEXT
from sqlalchemy.exc import SQLAlchemyError
from flask_wtf.csrf import CSRFProtect
from openpyxl.styles import Font

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@localhost/db-restructure'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'b35dfe6ce150230940bd145823034486'
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 150 * 1024 * 1024

csrf = CSRFProtect(app)

UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


def get_jakarta_time():
    """Get current time in Jakarta timezone"""
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    return datetime.now(jakarta_tz)


class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='Staff')

    tickets = db.relationship('Ticket', backref='pic_handle', lazy=True,
                              foreign_keys='Ticket.pic_handle_id')

    def __repr__(self):
        return f'<User {self.email}>'


class Product(db.Model):
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    complaint_type = db.Column(db.String(100), nullable=False)
    complaint_detail = db.Column(db.Text, nullable=False)
    status = db.Column(db.Integer, nullable=False, default=1)

    tickets = db.relationship(
        'Ticket', backref='product', lazy=True, foreign_keys='Ticket.product_id')

    def __repr__(self):
        return f'<Product {self.complaint_type}>'

class CollectionTemplate(db.Model):
    __tablename__ = 'collection_templates'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    detail = db.Column(LONGTEXT, nullable=False)
    status = db.Column(db.Integer, nullable=False, default=1)
    created_at = db.Column(db.DateTime, nullable=False,
                           default=lambda: get_jakarta_time())
    updated_at = db.Column(db.DateTime, nullable=False,
                           default=lambda: get_jakarta_time(),
                           onupdate=lambda: get_jakarta_time())

    def __repr__(self):
        return f'<CollectionTemplate {self.title}>'

class ReplyTemplate(db.Model):
    __tablename__ = 'reply_templates'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    detail = db.Column(LONGTEXT, nullable=False)
    status = db.Column(db.Integer, nullable=False, default=1)
    created_at = db.Column(db.DateTime, nullable=False,
                           default=lambda: get_jakarta_time())
    updated_at = db.Column(db.DateTime, nullable=False,
                           default=lambda: get_jakarta_time(),
                           onupdate=lambda: get_jakarta_time())

    def __repr__(self):
        return f'<ReplyTemplate {self.title}>'

class Ticket(db.Model):
    __tablename__ = 'tickets'
    id = db.Column(db.Integer, primary_key=True)
    no_ticket = db.Column(db.String(50), unique=True, nullable=False)
    tanggal_pengaduan = db.Column(
        db.DateTime, nullable=False, default=lambda: get_jakarta_time())
    tanggal_pengerjaan = db.Column(
        db.DateTime, nullable=False, default=lambda: get_jakarta_time())
    nama = db.Column(db.String(100), nullable=False)
    phone_pengajuan = db.Column(db.String(20), nullable=False)
    phone_aktif = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    order_number = db.Column(db.String(500), nullable=True)
    nominal_order = db.Column(db.String(500), nullable=True)
    pic_handle_id = db.Column(
        db.Integer, db.ForeignKey('users.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey(
        'products.id'), nullable=False)
    kanal_pengaduan = db.Column(db.String(50), nullable=False)
    detail_problem = db.Column(db.Text, nullable=False)
    tipe_pengaduan = db.Column(db.String(100), nullable=False)
    detail_pengaduan = db.Column(db.Text, nullable=False)
    user_respon = db.Column(db.Text, nullable=True)
    note = db.Column(db.Text, nullable=True)
    status = db.Column(db.Integer, nullable=False, default=1)
    case_progress = db.Column(db.Integer, nullable=True, default=1)
    tanggal_tutup_aduan = db.Column(db.DateTime, nullable=True)
    close_date = db.Column(db.DateTime, nullable=True)
    close_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    close_user = db.relationship('User', foreign_keys=[close_by])
    move_date = db.Column(db.DateTime, nullable=True)
    move_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    move_user = db.relationship('User', foreign_keys=[move_by])
    sla = db.Column(db.Integer, nullable=True, default=10)

    def __repr__(self):
        return f'<Ticket {self.no_ticket}>'


class Document(db.Model):
    __tablename__ = 'documents'

    id = db.Column(db.Integer, primary_key=True)
    document_name = db.Column(db.String(100), nullable=False)
    status = db.Column(db.Integer, default=1)


class File(db.Model):
    __tablename__ = 'files'

    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey(
        'tickets.id'), nullable=False)
    document_id = db.Column(db.Integer, db.ForeignKey(
        'documents.id'), nullable=False)
    file_uploaded = db.Column(db.String(255), nullable=False)
    uploaded_at = db.Column(db.DateTime, nullable=False,
                            default=datetime.utcnow)
    uploaded_by = db.Column(
        db.Integer, db.ForeignKey('users.id'), nullable=False)

    document = db.relationship(
        'Document', backref=db.backref('files', lazy=True))
    user = db.relationship(
        'User', backref=db.backref('uploaded_files', lazy=True))
    ticket = db.relationship('Ticket', backref=db.backref('files', lazy=True))

    def __repr__(self):
        return f'<File {self.file_uploaded}>'


class Notes(db.Model):
    __tablename__ = 'notes'
    id = db.Column(db.Integer, primary_key=True)
    send_by_id = db.Column(
        db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False,
                           default=lambda: get_jakarta_time())
    ticket_id = db.Column(db.Integer, db.ForeignKey(
        'tickets.id'), nullable=True)
    content = db.Column(db.Text, nullable=False)
    type_note = db.Column(db.String(50), nullable=False)

    sender = db.relationship('User', backref='notes_sent')
    ticket = db.relationship('Ticket', backref='notes')

    def __repr__(self):
        return f'<Notes {self.id} for Ticket {self.ticket_id if self.ticket_id else "N/A"}>'


class Docs1(db.Model):
    __tablename__ = 'docs1'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    upload_date = db.Column(db.DateTime, nullable=False,
                            default=lambda: get_jakarta_time())
    upload_by_id = db.Column(
        db.Integer, db.ForeignKey('users.id'), nullable=False)
    ticket_id = db.Column(db.Integer, db.ForeignKey(
        'tickets.id'), nullable=False)

    uploader = db.relationship('User', backref='docs1_uploaded')
    ticket = db.relationship('Ticket', backref='docs1_files')

    def __repr__(self):
        return f'<Docs1 {self.filename} for Ticket {self.ticket_id}>'


class Docs2(db.Model):
    __tablename__ = 'docs2'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    upload_date = db.Column(db.DateTime, nullable=False,
                            default=lambda: get_jakarta_time())
    upload_by_id = db.Column(
        db.Integer, db.ForeignKey('users.id'), nullable=False)
    ticket_id = db.Column(db.Integer, db.ForeignKey(
        'tickets.id'), nullable=False)

    uploader = db.relationship('User', backref='docs2_uploaded')
    ticket = db.relationship('Ticket', backref='docs2_files')

    def __repr__(self):
        return f'<Docs2 {self.filename} for Ticket {self.ticket_id}>'


class Tenor(db.Model):
    __tablename__ = 'tenors'
    id = db.Column(db.Integer, primary_key=True)
    nomor_kontrak = db.Column(db.String(500), nullable=False)
    tenor_1 = db.Column(db.Integer, nullable=True)
    nominal_tenor_1 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_1 = db.Column(db.Date, nullable=True)

    tenor_2 = db.Column(db.Integer, nullable=True)
    nominal_tenor_2 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_2 = db.Column(db.Date, nullable=True)

    tenor_3 = db.Column(db.Integer, nullable=True)
    nominal_tenor_3 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_3 = db.Column(db.Date, nullable=True)

    tenor_4 = db.Column(db.Integer, nullable=True)
    nominal_tenor_4 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_4 = db.Column(db.Date, nullable=True)

    tenor_5 = db.Column(db.Integer, nullable=True)
    nominal_tenor_5 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_5 = db.Column(db.Date, nullable=True)

    tenor_6 = db.Column(db.Integer, nullable=True)
    nominal_tenor_6 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_6 = db.Column(db.Date, nullable=True)

    tenor_7 = db.Column(db.Integer, nullable=True)
    nominal_tenor_7 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_7 = db.Column(db.Date, nullable=True)

    tenor_8 = db.Column(db.Integer, nullable=True)
    nominal_tenor_8 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_8 = db.Column(db.Date, nullable=True)

    tenor_9 = db.Column(db.Integer, nullable=True)
    nominal_tenor_9 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_9 = db.Column(db.Date, nullable=True)

    tenor_10 = db.Column(db.Integer, nullable=True)
    nominal_tenor_10 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_10 = db.Column(db.Date, nullable=True)

    tenor_11 = db.Column(db.Integer, nullable=True)
    nominal_tenor_11 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_11 = db.Column(db.Date, nullable=True)

    tenor_12 = db.Column(db.Integer, nullable=True)
    nominal_tenor_12 = db.Column(db.Numeric(
        precision=15, scale=2), nullable=True)
    ovd_12 = db.Column(db.Date, nullable=True)

    total_nominal = db.Column(db.Numeric(
        precision=15, scale=2), nullable=False)

    total_nominal_akhir = db.Column(db.Numeric(
        precision=15, scale=2), nullable=False)

    ticket_id = db.Column(db.Integer, db.ForeignKey(
        'tickets.id'), nullable=False)
    ticket = db.relationship('Ticket', backref='tenors')

    def __repr__(self):
        return f'<Tenor {self.nomor_kontrak} for Ticket {self.ticket_id}>'

# class Message(db.Model):
#     __tablename__ = 'messages'
#     id = db.Column(db.Integer, primary_key=True)
#     sender_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
#     receiver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
#     content = db.Column(db.Text, nullable=False)
#     timestamp = db.Column(db.DateTime, default=db.func.current_timestamp())

#     sender = db.relationship('User', foreign_keys=[sender_id], backref='sent_messages')
#     receiver = db.relationship('User', foreign_keys=[receiver_id], backref='received_messages')

#     def __repr__(self):
#         return f"<Message from {self.sender_id} to {self.receiver_id}>"


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def _recalc_totals(tenor_obj):
    sisa = Decimal('0')
    sisa_count = 0
    lunas_count = 0
    for i in range(1, 12 + 1):
        val = getattr(tenor_obj, f'nominal_tenor_{i}', None) or Decimal('0')
        if val > 0:
            sisa += val
            sisa_count += 1
        else:
            lunas_count += 1
    tenor_obj.total_nominal_akhir = sisa
    return sisa, sisa_count, lunas_count


def generate_staff_id():
    last_user = User.query.filter(User.staff_id.like('UATAS%')).order_by(
        db.func.cast(db.func.substring(User.staff_id, 6), db.Integer).desc()).first()
    if last_user:
        last_num = int(last_user.staff_id[5:])
        new_num = last_num + 1
    else:
        new_num = 1
    return f'UATAS{new_num}'


def generate_ticket_number():
    """Generate a unique ticket number in format UATASddmmyyNNN"""
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    now = datetime.now(jakarta_tz)
    date_str = now.strftime('%d%m%y')

    pattern = f"^UATAS{date_str}\\d{{3}}$"

    last_ticket_today = Ticket.query.filter(
        Ticket.no_ticket.op('REGEXP')(pattern)
    ).order_by(
        db.func.cast(
            db.func.substring(Ticket.no_ticket, -3), db.Integer
        ).desc()
    ).first()

    if last_ticket_today:
        last_num = int(last_ticket_today.no_ticket[-3:])
        new_num = last_num + 1
    else:
        new_num = 1

    return f"UATAS{date_str}{new_num:03d}"

def reduce_sla_daily():
    with app.app_context():
        print(f"Memproses pengurangan SLA otomatis pada {get_jakarta_time()}")
        try:
            tickets_to_update = Ticket.query.filter(Ticket.sla > 0).all()

            for ticket in tickets_to_update:
                ticket.sla -= 1

            db.session.commit()
            print(
                f"Pengurangan SLA selesai. {len(tickets_to_update)} tiket diperbarui.")
        except Exception as e:
            db.session.rollback()
            print(f"Error saat mengurangi SLA: {e}")


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg',
                          'gif', 'doc', 'docx', 'xls', 'xlsx', 'zip', 'rar'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

##### Export to Google Sheets #####


def export_to_google_sheet():
    SCOPE = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    CREDS = Credentials.from_service_account_file(
        "precise-works-478204-i9-0c94fde3a08d.json",
        scopes=SCOPE
    )
    gc = gspread.authorize(CREDS)

    spreadsheet = gc.open("Report-spreadsheet")
    worksheet = spreadsheet.worksheet("Sheet1")

    data = []

    tenors = Tenor.query.join(Ticket).all()

    no = 1
    for tenor in tenors:
        t = tenor.ticket
        row = build_row(no, t, tenor)
        data.append(row)
        no += 1

    worksheet.update("A3", data)
    print("Export ke Google Sheet Sukses!")


locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')


def format_rupiah(value):
    if value is None or value == "":
        return ""
    if str(value) == "0.00":
        return 0

    angka = float(value)
    return "Rp " + "{:,.0f}".format(angka).replace(",", ".")


def format_tanggal(db_value):
    if not db_value:
        return ""
    try:
        if isinstance(db_value, str):
            dt = datetime.strptime(db_value, "%Y-%m-%d")
        else:
            dt = db_value
        return dt.strftime("%d/%m/%Y")
    except:
        return ""


def build_row(no, t, tenor):

    order_value = tenor.nomor_kontrak if tenor.nomor_kontrak else ""

    tenor_count = sum(
        1 for i in range(1, 13)
        if getattr(tenor, f"tenor_{i}")
    )

    row = [
        no,
        t.tanggal_pengaduan.strftime(
            "%Y-%m-%d") if t.tanggal_pengaduan else "",
        t.pic_handle.name,
        t.email,
        t.nama,
        t.phone_pengajuan,
        "VPN",
        order_value,
        tenor_count,
    ]

    for i in range(1, 13):

        nominal_raw = getattr(tenor, f"nominal_tenor_{i}")
        nominal = format_rupiah(nominal_raw)

        jt_raw = getattr(tenor, f"ovd_{i}")
        jt = format_tanggal(jt_raw)

        row.append(nominal)
        row.append(jt)

    return row


@app.route("/export-google-sheet")
def export_google_sheet_route():
    try:
        export_to_google_sheet()
        flash("Export ke Google Sheet berhasil!", "success")
    except Exception as e:
        flash(f"Export gagal: {str(e)}", "danger")

    return redirect(url_for("dashboard"))

##### Export to Google Sheets #####


@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()

        if user and check_password_hash(user.password, password):
            login_user(user)
            flash(f'Login berhasil sebagai {user.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Email atau password salah.', 'danger')
            return redirect(url_for('login'))

    return render_template('login.html')


jakarta_tz = pytz.timezone('Asia/Jakarta')
now_id = datetime.now(jakarta_tz)
current_year = now_id.year


@app.route('/dashboard')
@login_required
def dashboard():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    status = request.args.get('status', '', type=str)

    if current_user.role == 'Admin':
        query = Ticket.query

        if search:
            query = query.filter(
                db.or_(
                    Ticket.no_ticket.ilike(f"%{search}%"),
                    Ticket.nama.ilike(f"%{search}%"),
                    Ticket.email.ilike(f"%{search}%"),
                    Ticket.phone_aktif.ilike(f"%{search}%")
                )
            )

        if status:
            try:
                status_value = int(status)
                query = query.filter(Ticket.status == status_value)
            except ValueError:
                pass

        total_tickets_status = query.filter(
            Ticket.status.in_([1, 2, 3])).count()
        status_1 = query.filter_by(status=1).count()
        status_2 = query.filter_by(status=2).count()
        status_3 = query.filter_by(status=3).count()

        if total_tickets_status > 0:
            percent_1 = round((status_1 / total_tickets_status) * 100, 2)
            percent_2 = round((status_2 / total_tickets_status) * 100, 2)
            percent_3 = round((status_3 / total_tickets_status) * 100, 2)
        else:
            percent_1 = percent_2 = percent_3 = 0

        tickets_paginated = query.order_by(Ticket.id.desc()).paginate(
            page=page, per_page=5, error_out=False
        )

        raw_upcoming_tickets = Ticket.query.filter(
            Ticket.sla.between(1, 3)
        ).order_by(Ticket.tanggal_pengaduan.asc()).limit(5).all()

        jakarta_tz = pytz.timezone('Asia/Jakarta')
        current_date = datetime.now(jakarta_tz)

        upcoming_tickets_with_sla_end = []
        for ticket in raw_upcoming_tickets:
            sla_end_date = current_date + timedelta(days=ticket.sla)
            upcoming_tickets_with_sla_end.append({
                'ticket': ticket,
                'sla_end_date': sla_end_date
            })

        rows = (
            db.session.query(
                extract('month', Ticket.tanggal_pengaduan).label('bulan'),
                func.count(Ticket.id).label('jumlah')
            )
            .filter(extract('year', Ticket.tanggal_pengaduan) == current_year)
            .group_by('bulan')
            .order_by('bulan')
            .all()
        )

        month_counts = [0] * 12
        for bulan, jumlah in rows:
            month_counts[int(bulan) - 1] = int(jumlah)

        month_labels = ["Jan", "Feb", "Mar", "Apr", "May",
                        "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        products = Product.query.all()

        return render_template(
            'dashboard_admin.html',
            user=current_user,
            tickets=tickets_paginated,
            search=search,
            status=status,
            total_tickets_status=total_tickets_status,
            status_1=status_1,
            status_2=status_2,
            status_3=status_3,
            percent_1=percent_1,
            percent_2=percent_2,
            percent_3=percent_3,
            upcoming_tickets=upcoming_tickets_with_sla_end,
            month_labels=month_labels,
            month_counts=month_counts,
            current_year=current_year,
            products=products
        )

    elif current_user.role == 'Staff':
        query = Ticket.query

        if search:
            query = query.filter(
                or_(
                    Ticket.no_ticket.ilike(f"%{search}%"),
                    Ticket.nama.ilike(f"%{search}%"),
                    Ticket.email.ilike(f"%{search}%"),
                    Ticket.phone_aktif.ilike(f"%{search}%")
                )
            )

        if status:
            try:
                status_value = int(status)
                query = query.filter(Ticket.status == status_value)
            except ValueError:
                pass

        now = datetime.now()
        bulan_ini = now.month
        tahun_ini = now.year
        nama_bulan = f"{calendar.month_name[bulan_ini]} {tahun_ini}"

        total_case = (
            db.session.query(func.count(Ticket.id))
            .scalar()
        )

        status_counts = (
            db.session.query(Ticket.status, func.count(Ticket.id))
            .filter(
                Ticket.status.in_([1, 2, 3])
            )
            .group_by(Ticket.status)
            .all()
        )

        status_dict = {s: j for s, j in status_counts}
        status_1 = status_dict.get(1, 0)
        status_2 = status_dict.get(2, 0)
        status_3 = status_dict.get(3, 0)

        case_data = (
            db.session.query(
                func.ceil(
                    (extract('day', Ticket.tanggal_pengaduan) / 7)).label("minggu"),
                func.count(Ticket.id).label("jumlah")
            )
            .filter(
                Ticket.pic_handle_id == current_user.id,
                extract('month', Ticket.tanggal_pengaduan) == bulan_ini,
                extract('year', Ticket.tanggal_pengaduan) == tahun_ini
            )
            .group_by("minggu")
            .order_by("minggu")
            .all()
        )

        minggu_labels = []
        jumlah_case = []
        for c in case_data:
            start_day = int((c.minggu - 1) * 7 + 1)
            end_day = int(min(c.minggu * 7, 31))
            minggu_labels.append(f"{start_day}-{end_day}")
            jumlah_case.append(c.jumlah)

        tickets_paginated = query.order_by(Ticket.id.desc()).paginate(
            page=page, per_page=5, error_out=False)

        products = Product.query.all()

        return render_template(
            'dashboard_user.html',
            user=current_user,
            tickets=tickets_paginated,
            search=search,
            status=status,
            minggu_labels=minggu_labels,
            jumlah_case=jumlah_case,
            nama_bulan=nama_bulan,
            total_case=total_case,
            status_1=status_1,
            status_2=status_2,
            status_3=status_3,
            products=products
        )

    else:
        flash('Role tidak dikenali.', 'danger')
        return redirect(url_for('login'))


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('login'))

##### User Management Routes #####


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']

        if role not in ['Admin', 'Staff']:
            flash('Role tidak valid.', 'danger')
            return redirect(url_for('register'))

        existing = User.query.filter_by(email=email).first()
        if existing:
            flash('Email sudah terdaftar.', 'danger')
            return redirect(url_for('register'))

        hashed_pw = generate_password_hash(password)
        new_user = User(name=name, phone=phone, email=email,
                        password=hashed_pw, role=role)
        db.session.add(new_user)
        db.session.commit()
        flash('Registrasi berhasil! Silakan login.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/list-users')
@login_required
def list_users():
    if current_user.role != 'Admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard'))

    page = request.args.get('page', 1, type=int)
    per_page = 10

    users_pagination = User.query.paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template('user_list.html', users=users_pagination.items, pagination=users_pagination, user=current_user)


@app.route('/add-user', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'Admin':
        flash('Akses ditolak. Hanya Admin yang dapat menambah user.', 'danger')
        return redirect(url_for('list_users'))

    name = request.form.get('employeeInput')
    phone = request.form.get('phoneNumberInput')
    email = request.form.get('emailInput')
    password = request.form.get('passwordInput')
    role = request.form.get('role', 'Staff')

    if not all([name, phone, email, password]):
        flash('Semua field kecuali role wajib diisi.', 'danger')
        return redirect(url_for('list_users'))

    if role not in ['Admin', 'Staff']:
        flash('Role tidak valid. Gunakan Admin atau Staff.', 'danger')
        return redirect(url_for('list_users'))

    existing = User.query.filter_by(email=email).first()
    if existing:
        flash('Email sudah terdaftar.', 'danger')
        return redirect(url_for('list_users'))

    staff_id = generate_staff_id()

    hashed_pw = generate_password_hash(password)
    new_user = User(staff_id=staff_id, name=name, phone=phone,
                    email=email, password=hashed_pw, role=role)
    db.session.add(new_user)
    db.session.commit()
    flash(f'User berhasil ditambahkan! Staff ID: {staff_id}', 'success')
    return redirect(url_for('list_users'))


@app.route('/delete-user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'Admin':
        return jsonify({'message': 'Akses ditolak. Hanya Admin yang dapat menghapus pengguna.'}), 403

    user_to_delete = User.query.get_or_404(user_id)

    if user_to_delete.id == current_user.id:
        return jsonify({'message': 'Anda tidak dapat menghapus akun Anda sendiri.'}), 400

    try:
        db.session.delete(user_to_delete)
        db.session.commit()
        return jsonify({'message': f'User "{user_to_delete.name}" (ID: {user_to_delete.staff_id}) berhasil dihapus.'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Gagal menghapus user: {str(e)}'}), 500


@app.route('/update-user/<int:user_id>', methods=['POST'])
@login_required
def update_user(user_id):
    """
    Handles the update of a user's data via POST request.
    Expects user_id as a parameter and form data for the fields to update.
    """
    if current_user.role != 'Admin':
        return jsonify({'message': 'Akses ditolak. Hanya Admin yang dapat mengupdate pengguna.'}), 403

    user_to_update = User.query.get_or_404(user_id)

    data = request.get_json()
    if not data:
        return jsonify({'message': 'Data tidak ditemukan dalam permintaan.'}), 400

    new_name = data.get('name')
    new_email = data.get('email')
    new_phone = data.get('phone')
    new_role = data.get('role')
    new_password = data.get('password')

    if not new_name or not new_email or not new_phone or not new_role:
        return jsonify({'message': 'Semua field wajib diisi (kecuali password).'}), 400

    try:
        user_to_update.name = new_name
        user_to_update.email = new_email
        user_to_update.phone = new_phone
        user_to_update.role = new_role

        if new_password:
            user_to_update.password = generate_password_hash(new_password)

        db.session.commit()
        return jsonify({'message': f'Data user "{user_to_update.name}" (ID: {user_to_update.staff_id}) berhasil diperbarui.'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Gagal memperbarui user: {str(e)}'}), 500


@app.route('/upload_file/<int:document_id>/<int:ticket_id>', methods=['POST'])
@login_required
def upload_file(document_id, ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)

    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        flash("Tidak ada file yang dipilih untuk diupload.", "danger")
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    for file in files:
        if file.filename == '':
            continue

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join('static/uploads', filename)
            file.save(filepath)

            new_file = File(
                document_id=document_id,
                ticket_id=ticket_id,
                file_uploaded=filename,
                uploaded_by=current_user.id
            )
            db.session.add(new_file)
        else:
            flash(f'Tipe file tidak diizinkan: {file.filename}', 'danger')
            db.session.rollback()
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    ticket.status = 2

    try:
        db.session.commit()
        flash("File berhasil diupload!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Gagal mengupload file: {str(e)}", "danger")
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


@app.route('/view_file/<int:document_id>')
def view_file(document_id):
    from flask import send_from_directory, flash, redirect, url_for
    file_record = File.query.filter_by(document_id=document_id).first()

    if not file_record:
        flash('File belum diupload untuk dokumen ini.', 'warning')
        return redirect(url_for('upload_file'))

    return send_from_directory('static/uploads', file_record.file_uploaded)

##### Admin Management Routes #####


@app.route('/product-management')
@login_required
def product_management():
    if current_user.role != 'Admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard'))

    products = Product.query.all()

    return render_template('product_management.html', user=current_user, products=products)


@app.route('/document-management')
@login_required
def document_management():
    if current_user.role != 'Admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard'))

    documents = Document.query.all()
    return render_template('document_management.html', user=current_user, documents=documents)


@app.route('/add_document', methods=['POST'])
@login_required
def add_document():
    if current_user.role != 'Admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('document_management'))

    name = request.form.get('document_name')
    status = request.form.get('status', 1)

    new_doc = Document(document_name=name, status=status)
    db.session.add(new_doc)
    db.session.commit()
    flash('Document Category added successfully!', 'success')
    return redirect(url_for('document_management'))


@app.route('/update_document/<int:document_id>', methods=['POST'])
@login_required
def update_document(document_id):
    doc = Document.query.get_or_404(document_id)
    doc.status = request.form.get('status')
    db.session.commit()
    flash('Document status updated!', 'success')
    return redirect(url_for('document_management'))


@app.route('/delete_document/<int:document_id>', methods=['POST'])
@login_required
def delete_document(document_id):
    doc = Document.query.get_or_404(document_id)
    db.session.delete(doc)
    db.session.commit()
    flash('Document deleted successfully!', 'success')
    return redirect(url_for('document_management'))


@app.route('/create-product', methods=['GET', 'POST'])
@login_required
def create_product():
    """Create a new product/complaint type or add detail to existing complaint type"""
    if current_user.role != 'Admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        complaint_type = request.form.get('typeInput')
        complaint_detail = request.form.get('detailInput')
        status = request.form.get('status', 1)

        if not complaint_type or not complaint_detail:
            flash('Complaint Type dan Detail Complaint wajib diisi.', 'danger')
            return render_template('product_management.html', user=current_user)

        existing_product = Product.query.filter_by(
            complaint_type=complaint_type).first()

        if existing_product:
            existing_details = existing_product.complaint_detail.split(
                ',') if existing_product.complaint_detail else []
            existing_details = [detail.strip() for detail in existing_details]

            if complaint_detail.strip() not in existing_details:
                existing_details.append(complaint_detail.strip())
                existing_product.complaint_detail = ', '.join(existing_details)
                flash(
                    f'Detail complaint berhasil ditambahkan ke "{complaint_type}"!', 'success')
            else:
                flash(
                    f'Detail complaint "{complaint_detail}" sudah ada dalam "{complaint_type}".', 'warning')
        else:
            new_product = Product(
                complaint_type=complaint_type,
                complaint_detail=complaint_detail,
                status=int(status)
            )
            db.session.add(new_product)
            flash('Product berhasil ditambahkan!', 'success')

        try:
            db.session.commit()
            return redirect(url_for('product_management'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui product: {str(e)}', 'danger')
            return render_template('product_management.html', user=current_user)

    return render_template('product_management.html', user=current_user)


@app.route('/update-product-status/<int:product_id>', methods=['POST'])
@login_required
def update_product_status(product_id):
    """Update the status and/or complaint detail of a specific product"""
    if current_user.role != 'Admin':
        flash('Akses ditolak. Hanya Admin yang dapat memperbarui product.', 'danger')
        return redirect(url_for('product_management'))

    product = Product.query.get_or_404(product_id)

    new_status = request.form.get('status')
    new_detail = request.form.get('complaint_detail')

    if new_status is None and new_detail is None:
        flash('Tidak ada data yang diperbarui.', 'warning')
        return redirect(url_for('product_management'))

    try:
        if new_status is not None:
            product.status = int(new_status)
        if new_detail is not None:
            product.complaint_detail = new_detail.strip()

        db.session.commit()
        flash(
            f'Product "{product.complaint_type}" berhasil diperbarui.', 'success')
        return redirect(url_for('product_management'))
    except ValueError:
        db.session.rollback()
        flash('Status harus berupa angka.', 'danger')
        return redirect(url_for('product_management'))
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal memperbarui product: {str(e)}', 'danger')
        return redirect(url_for('product_management'))


@app.route('/delete-product/<int:product_id>', methods=['POST'])
@login_required
def delete_product(product_id):
    """Delete a specific product"""
    if current_user.role != 'Admin':
        flash('Akses ditolak. Hanya Admin yang dapat menghapus product.', 'danger')
        return redirect(url_for('product_management'))

    product = Product.query.get_or_404(product_id)

    try:
        db.session.delete(product)
        db.session.commit()
        flash(
            f'Product "{product.complaint_type}" berhasil dihapus.', 'success')
        return redirect(url_for('product_management'))
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus product: {str(e)}', 'danger')
        return redirect(url_for('product_management'))

##### Staff Management Routes #####


@app.route('/create-tickets', methods=['GET', 'POST'])
@login_required
def create_tickets():
    if current_user.role != 'Staff':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':

        tanggal_pengaduan_str = request.form.get('tanggalPengaduan')
        if not tanggal_pengaduan_str:
            flash('Tanggal pengaduan harus diisi.', 'danger')
            products_for_template = Product.query.all()
            return render_template('create_ticket.html', user=current_user, products=products_for_template)

        try:
            tanggal_pengaduan_obj = datetime.strptime(
                tanggal_pengaduan_str, '%Y-%m-%d')
            jakarta_tz = pytz.timezone('Asia/Jakarta')
            tanggal_pengaduan_obj = jakarta_tz.localize(tanggal_pengaduan_obj)
        except ValueError:
            flash('Format tanggal_pengaduan tidak valid. Gunakan YYYY-MM-DD.', 'danger')
            products_for_template = Product.query.all()
            return render_template('create_ticket.html', user=current_user, products=products_for_template)

        tanggal_pengerjaan_obj = datetime.now(jakarta_tz)
        no_ticket = generate_ticket_number()

        nama = request.form.get('namaNasabah')
        phone_pengajuan = request.form.get('phonePengajuan')
        phone_aktif = request.form.get('phoneAktif')
        email = request.form.get('emailNasabah')

        order_number_raw = request.form.get('orderNumber')
        order_number = None
        order_numbers_list = []

        if order_number_raw:
            order_numbers_list = [
                num.strip() for num in order_number_raw.split(',') if num.strip()]
            order_number = ",".join(order_numbers_list)

        nominal_list = request.form.getlist('nominalOrder[]')
        nominal_list = [n.strip() for n in nominal_list if n.strip()]

        nominal_order = ",".join(nominal_list)

        if len(order_numbers_list) != len(nominal_list):
            flash("Jumlah Nominal harus sama dengan jumlah Order Number.", "danger")
            products_for_template = Product.query.all()
            return render_template('create_ticket.html', user=current_user, products=products_for_template)

        pic_handle_id = current_user.id
        kanal_pengaduan = request.form.get('kanalPengaduan')
        status = request.form.get('statusCase')
        detail_problem = request.form.get('detailProblem')
        tipe_pengaduan = request.form.get('tipePengaduan')
        detail_pengaduan = request.form.get('detailPengaduan')

        if not all([nama, phone_pengajuan, phone_aktif, email, status, kanal_pengaduan, detail_problem, tipe_pengaduan, detail_pengaduan]):
            flash('Semua field wajib diisi.', 'danger')
            products_for_template = Product.query.all()
            return render_template('create_ticket.html', user=current_user, products=products_for_template)

        product = Product.query.filter(
            Product.complaint_type == tipe_pengaduan,
            Product.complaint_detail.contains(detail_pengaduan)
        ).first()

        if not product:
            product = Product.query.filter_by(
                complaint_type=tipe_pengaduan,
                complaint_detail=detail_pengaduan
            ).first()

        if not product:
            flash('Product tidak ditemukan untuk tipe & detail pengaduan.', 'danger')
            products_for_template = Product.query.all()
            return render_template('create_ticket.html', user=current_user, products=products_for_template)

        new_ticket = Ticket(
            no_ticket=no_ticket,
            tanggal_pengaduan=tanggal_pengaduan_obj,
            tanggal_pengerjaan=tanggal_pengerjaan_obj,
            nama=nama,
            phone_pengajuan=phone_pengajuan,
            phone_aktif=phone_aktif,
            email=email,
            order_number=order_number,
            nominal_order=nominal_order,
            pic_handle_id=pic_handle_id,
            product_id=product.id,
            status=status,
            kanal_pengaduan=kanal_pengaduan,
            detail_problem=detail_problem,
            tipe_pengaduan=tipe_pengaduan,
            detail_pengaduan=detail_pengaduan,
            sla=10
        )

        try:
            db.session.add(new_ticket)
            db.session.commit()
            flash(f'Ticket {new_ticket.no_ticket} berhasil dibuat!', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal membuat ticket: {str(e)}', 'danger')
            products_for_template = Product.query.all()
            return render_template('create_ticket.html', user=current_user, products=products_for_template)

    products = Product.query.all()
    grouped_details = {}

    for product in products:
        details_list = [detail.strip()
                        for detail in product.complaint_detail.split(',')]
        if product.complaint_type not in grouped_details:
            grouped_details[product.complaint_type] = []
        for detail in details_list:
            if detail not in grouped_details[product.complaint_type]:
                grouped_details[product.complaint_type].append(detail)

    return render_template(
        'create_ticket.html',
        user=current_user,
        products=products,
        grouped_details=grouped_details
    )


@app.route('/case-open', methods=['GET'])
@login_required
def case_open():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '', type=str)

    query = Ticket.query.filter_by(status=1, case_progress=1)

    if search_query:
        query = query.filter(
            db.or_(
                Ticket.no_ticket.contains(search_query),
                Ticket.nama.contains(search_query),
                Ticket.tipe_pengaduan.contains(search_query)
            )
        )

    pagination = query.paginate(
        page=page, per_page=10, error_out=False
    )
    open_tickets = pagination.items

    return render_template('case_open.html', user=current_user, tickets=open_tickets, pagination=pagination, search_query=search_query)


@app.route('/case-collection', methods=['GET'])
@login_required
def case_collection():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '', type=str)

    query = Ticket.query.filter_by(
        case_progress=2).filter(Ticket.status != '3')

    if search_query:
        query = query.filter(
            db.or_(
                Ticket.no_ticket.contains(search_query),
                Ticket.nama.contains(search_query)
            )
        )

    pagination = query.paginate(
        page=page, per_page=10, error_out=False
    )

    open_tickets = pagination.items

    return render_template('case_collection.html', user=current_user, tickets=open_tickets, pagination=pagination, search_query=search_query)


@app.route('/case-collection-close', methods=['GET'])
@login_required
def case_collection_close():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '', type=str)

    query = Ticket.query.filter_by(case_progress=2, status=3)

    if search_query:
        query = query.filter(
            db.or_(
                Ticket.no_ticket.contains(search_query),
                Ticket.nama.contains(search_query)
            )
        )

    pagination = query.paginate(
        page=page, per_page=10, error_out=False
    )

    open_tickets = pagination.items

    return render_template('case_collection_close.html', user=current_user, tickets=open_tickets, pagination=pagination, search_query=search_query)


@app.route('/case-detail-collection/<int:ticket_id>')
@login_required
def case_detail_collection(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    documents = Document.query.all()
    files = File.query.filter_by(ticket_id=ticket.id).all()
    notes = Notes.query.filter_by(ticket_id=ticket.id).order_by(
        Notes.created_at.desc()).all()
    docs1_files = Docs1.query.filter_by(ticket_id=ticket.id).all()
    docs2_files = Docs2.query.filter_by(ticket_id=ticket.id).all()
    tenors = Tenor.query.filter_by(ticket_id=ticket.id).all()

    tenor_by_kontrak = {(t.nomor_kontrak or '').strip(): t for t in tenors}

    order_numbers_list = [
        s.strip() for s in (ticket.order_number or '').split(',')
        if s.strip()
    ]

    nominal_list = [
        n.strip() for n in (ticket.nominal_order or '').split(',')
        if n.strip()
    ]

    def format_rupiah(number_str):
        try:
            amount = int(float(number_str))
            return f"Rp {amount:,.0f}".replace(",", ".")
        except:
            return number_str

    def is_nullish(v):
        if v is None:
            return True
        if isinstance(v, str) and v.strip().upper() in {"", "NULL", "(NULL)"}:
            return True
        return False

    kontrak_items = []

    for idx, onum in enumerate(order_numbers_list):

        nominal_value = nominal_list[idx] if idx < len(nominal_list) else None
        nominal_formatted = (
            format_rupiah(nominal_value) if nominal_value else None
        )

        t = tenor_by_kontrak.get(onum)
        if not t:
            continue

        slots_valid = []

        for i in range(1, 13):
            tenor_val = getattr(t, f'tenor_{i}', None)
            if is_nullish(tenor_val):
                continue

            nominal_val = getattr(t, f'nominal_tenor_{i}', None)
            nominal_num = float(nominal_val or 0)

            slots_valid.append({
                "no": i,
                "tenor": tenor_val,
                "nominal": nominal_num,
            })

        tenor_lunas_list = [
            {**s, "is_lunas": True}
            for s in slots_valid if s["nominal"] <= 0
        ]
        tenor_aktif_list = [
            s for s in slots_valid if s["nominal"] > 0
        ]

        lunas_count = len(tenor_lunas_list)
        sisa_count = len(tenor_aktif_list)

        nominal_satuan = next(
            (s["nominal"] for s in tenor_aktif_list if s["nominal"] > 0),
            0.0
        )

        total_nominal_akhir = float(t.total_nominal_akhir or 0)
        total_tenor_lunas_amount = lunas_count * nominal_satuan

        tenor_aktif_first = tenor_aktif_list[0] if tenor_aktif_list else None

        tenor_aktif_sisa_grouped = None
        if len(tenor_aktif_list) > 1:
            sisa = tenor_aktif_list[1:]
            tenor_aktif_sisa_grouped = {
                "no": f"{sisa[0]['no']}-{sisa[-1]['no']}",
                "tenor": f"Tenor {sisa[0]['no']}–{sisa[-1]['no']}",
                "nominal": sum(x["nominal"] for x in sisa),
                "is_grouped": True
            }

        kontrak_items.append({
            "order_number": onum,
            "nominal": nominal_formatted,
            "tenor_id": t.id,
            "tenor_lunas_list": tenor_lunas_list,
            "tenor_aktif_first": tenor_aktif_first,
            "tenor_aktif_sisa_grouped": tenor_aktif_sisa_grouped,
            "jumlah_tenor_aktif": len(tenor_aktif_list),

            "lunas_count": lunas_count,
            "sisa_count": sisa_count,
            "nominal_satuan": nominal_satuan,
            "total_nominal_akhir": total_nominal_akhir,
            "total_tenor_lunas_amount": total_tenor_lunas_amount,

            "total_nominal": float(t.total_nominal or 0),
            "total_nominal_formatted": format_rupiah(t.total_nominal),
        })

    return render_template(
        'case_detail_collection.html',
        ticket=ticket,
        user=current_user,
        documents=documents,
        files=files,
        notes=notes,
        docs1_files=docs1_files,
        docs2_files=docs2_files,
        kontrak_items=kontrak_items,
        order_numbers_list=order_numbers_list
    )

@app.route('/tenor/delete/<int:tenor_id>', methods=['POST'])
def delete_tenor(tenor_id):
    tenor = Tenor.query.get_or_404(tenor_id)

    try:
        db.session.delete(tenor)
        db.session.commit()
        flash('Tenor berhasil dihapus', 'success')
    except SQLAlchemyError:
        db.session.rollback()
        flash('Gagal menghapus tenor', 'danger')

    return redirect(request.referrer or url_for('case_detail_collection'))

@app.route('/ticket/<int:ticket_id>/order/delete', methods=['POST'])
@login_required
def delete_order_number(ticket_id):
    order_number = request.form.get('order_number')

    ticket = Ticket.query.get_or_404(ticket_id)

    orders = [s.strip() for s in (ticket.order_number or '').split(',') if s.strip()]
    nominals = [s.strip() for s in (ticket.nominal_order or '').split(',') if s.strip()]

    if order_number in orders:
        idx = orders.index(order_number)
        orders.pop(idx)
        if idx < len(nominals):
            nominals.pop(idx)

        ticket.order_number = ','.join(orders)
        ticket.nominal_order = ','.join(nominals)
        db.session.commit()

        flash('Order berhasil dihapus', 'success')
    else:
        flash('Order tidak ditemukan', 'danger')

    return redirect(url_for('ticket_detail', ticket_id=ticket_id))

@app.route("/tenor/lunas/<int:tenor_id>/<int:no>", methods=["POST"])
@login_required
def tenor_lunas(tenor_id, no):
    if not (1 <= no <= 12):
        return jsonify({"ok": False, "error": "No tenor tidak valid"}), 400

    t = Tenor.query.get_or_404(tenor_id)
    field = f"nominal_tenor_{no}"
    if not hasattr(t, field):
        return jsonify({"ok": False, "error": "Field tenor tidak ditemukan"}), 400

    setattr(t, field, Decimal('0'))
    total_sisa, sisa_count, lunas_count = _recalc_totals(t)
    db.session.commit()

    kontrak = (t.nomor_kontrak or '').strip()
    flash(f"Tenor {no} pada kontrak {kontrak} berhasil dilunasi.", "info")

    return jsonify({
        "ok": True,
        "tenor_id": t.id,
        "no": no,
        "total_nominal_akhir": int(total_sisa),
        "sisa_count": sisa_count,
        "lunas_count": lunas_count,
        "redirect_url": url_for("case_detail_collection", ticket_id=t.ticket_id)
    })


@app.route('/case-process', methods=['GET'])
@login_required
def case_process():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '', type=str)

    query = Ticket.query.filter_by(status=2, case_progress=1)

    if search_query:
        query = query.filter(
            db.or_(
                Ticket.no_ticket.contains(search_query),
                Ticket.nama.contains(search_query),
                Ticket.tipe_pengaduan.contains(search_query)
            )
        )

    pagination = query.paginate(
        page=page, per_page=10, error_out=False
    )
    open_tickets = pagination.items

    return render_template('case_process.html', user=current_user, tickets=open_tickets, pagination=pagination, search_query=search_query)


@app.route('/case-close', methods=['GET'])
@login_required
def case_close():
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '', type=str)

    query = Ticket.query.filter_by(status=3, case_progress=1)

    if search_query:
        query = query.filter(
            db.or_(
                Ticket.no_ticket.contains(search_query),
                Ticket.nama.contains(search_query),
                Ticket.tipe_pengaduan.contains(search_query)
            )
        )

    pagination = query.paginate(
        page=page, per_page=10, error_out=False
    )
    open_tickets = pagination.items

    return render_template('case_close.html', user=current_user, tickets=open_tickets, pagination=pagination, search_query=search_query)


@app.route('/ticket/<int:ticket_id>')
@login_required
def ticket_detail(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    documents = Document.query.all()
    files = File.query.filter_by(ticket_id=ticket.id).all()
    notes = Notes.query.filter_by(ticket_id=ticket.id).order_by(
        Notes.created_at.desc()).all()
    docs1_files = Docs1.query.filter_by(ticket_id=ticket.id).all()
    docs2_files = Docs2.query.filter_by(ticket_id=ticket.id).all()

    tenors = Tenor.query.filter_by(ticket_id=ticket.id).all()
    tenor_by_kontrak = {(t.nomor_kontrak or '').strip(): t for t in tenors}

    order_numbers_list = [s.strip() for s in (
        ticket.order_number or '').split(',') if s.strip()]

    nominal_list = [n.strip() for n in (
        ticket.nominal_order or '').split(',') if n.strip()]

    def format_rupiah(number_str):
        try:
            amount = int(number_str)
            return f"Rp {amount:,.0f}".replace(",", ".")
        except:
            return number_str

    kontrak_items = []
    for idx, onum in enumerate(order_numbers_list):
        t = tenor_by_kontrak.get(onum)

        nominal_value = nominal_list[idx] if idx < len(nominal_list) else None
        nominal_formatted = format_rupiah(
            nominal_value) if nominal_value else None

        kontrak_items.append({
            'order_number': onum,
            'nominal': nominal_formatted,
            'tenor_id': t.id if t else None
        })

    return render_template(
        'case_detail.html',
        ticket=ticket,
        user=current_user,
        documents=documents,
        files=files,
        notes=notes,
        docs1_files=docs1_files,
        docs2_files=docs2_files,
        tenors=tenors,
        order_numbers_list=order_numbers_list,
        kontrak_items=kontrak_items
    )

@app.route('/ticket/<int:ticket_id>/add_order_number', methods=['POST'])
@login_required
def add_order_number(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    new_order_number = request.form.get('order_number', '').strip()

    if not new_order_number:
        flash('Order Number tidak boleh kosong', 'danger')
        return redirect(request.referrer or url_for('ticket_detail', ticket_id=ticket_id))

    existing_order_numbers = [s.strip() for s in (ticket.order_number or '').split(',') if s.strip()]

    if new_order_number not in existing_order_numbers:
        existing_order_numbers.append(new_order_number)
        ticket.order_number = ','.join(existing_order_numbers)
        try:
            db.session.commit()
            flash(f'Order Number {new_order_number} berhasil ditambahkan', 'success')
        except:
            db.session.rollback()
            flash('Gagal menambahkan Order Number', 'danger')
    else:
        flash('Order Number sudah ada', 'warning')

    return redirect(request.referrer or url_for('ticket_detail', ticket_id=ticket_id))

@app.route('/add-note/<int:ticket_id>', methods=['POST'])
@login_required
def add_note_to_ticket(ticket_id):
    """Tambah notes baru untuk ticket tertentu"""
    ticket = Ticket.query.get_or_404(ticket_id)
    content = request.form.get('content')
    type_note = request.form.get('type_note')

    if not content:
        flash('Isi notes tidak boleh kosong.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    new_note = Notes(
        send_by_id=current_user.id,
        ticket_id=ticket.id,
        content=content,
        type_note=type_note
    )

    try:
        db.session.add(new_note)
        db.session.commit()
        flash('Notes berhasil ditambahkan.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menambahkan notes: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('ticket_detail', ticket_id=ticket_id))


@app.route('/upload-docs1/<int:ticket_id>', methods=['POST'])
@login_required
def upload_docs1(ticket_id):
    """Upload file ke docs1 untuk ticket tertentu, menggantikan yang lama jika ada"""
    ticket = Ticket.query.get_or_404(ticket_id)

    if 'file' not in request.files:
        flash('Tidak ada file yang dipilih.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    file = request.files['file']
    manual_filename = request.form.get('manual_filename', '').strip()

    if file.filename == '':
        flash('Tidak ada file yang dipilih.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    if file and allowed_file(file.filename):
        original_filename = secure_filename(file.filename)
        original_name_base, original_extension = os.path.splitext(
            original_filename)

        if manual_filename:
            name_to_save_base = secure_filename(manual_filename)
        else:
            name_to_save_base = original_name_base

        filename_to_save = name_to_save_base + original_extension

        if not filename_to_save:
            flash('Nama file tidak valid.', 'danger')
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))

        existing_file = Docs1.query.filter_by(ticket_id=ticket.id).first()
        if existing_file:
            if os.path.exists(existing_file.file_path):
                os.remove(existing_file.file_path)
            db.session.delete(existing_file)

        relative_path_from_static = os.path.join(
            'uploads', 'docs1', f"{ticket.no_ticket}_{filename_to_save}").replace('\\', '/')
        full_file_path = os.path.join(
            app.root_path, 'static', relative_path_from_static)

        os.makedirs(os.path.dirname(full_file_path), exist_ok=True)

        file.save(full_file_path)

        new_file = Docs1(
            filename=filename_to_save,
            file_path=relative_path_from_static,
            upload_by_id=current_user.id,
            ticket_id=ticket.id
        )

        try:
            db.session.add(new_file)
            db.session.commit()
            flash(
                f'File {filename_to_save} berhasil diupload ke Restructure Docs (menggantikan file sebelumnya jika ada).', 'success')
        except Exception as e:
            db.session.rollback()
            if os.path.exists(full_file_path):
                os.remove(full_file_path)
            flash(f'Gagal mengupload file: {str(e)}', 'danger')
    else:
        flash('Tipe file tidak diizinkan.', 'danger')

    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


@app.route('/upload-docs2/<int:ticket_id>', methods=['POST'])
@login_required
def upload_docs2(ticket_id):
    """Upload file ke docs2 untuk ticket tertentu, menggantikan yang lama jika ada"""
    ticket = Ticket.query.get_or_404(ticket_id)

    if 'file' not in request.files:
        flash('Tidak ada file yang dipilih.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    file = request.files['file']
    manual_filename = request.form.get('manual_filename', '').strip()

    if file.filename == '':
        flash('Tidak ada file yang dipilih.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    if file and allowed_file(file.filename):
        original_filename = secure_filename(file.filename)
        original_name_base, original_extension = os.path.splitext(
            original_filename)

        if manual_filename:
            name_to_save_base = secure_filename(manual_filename)
        else:
            name_to_save_base = original_name_base

        filename_to_save = name_to_save_base + original_extension

        if not filename_to_save:
            flash('Nama file tidak valid.', 'danger')
            return redirect(url_for('ticket_detail', ticket_id=ticket_id))

        existing_file = Docs2.query.filter_by(ticket_id=ticket.id).first()
        if existing_file:
            if os.path.exists(existing_file.file_path):
                os.remove(existing_file.file_path)
            db.session.delete(existing_file)

        relative_path_from_static = os.path.join(
            'uploads', 'docs2', f"{ticket.no_ticket}_{filename_to_save}").replace('\\', '/')
        full_file_path = os.path.join(
            app.root_path, 'static', relative_path_from_static)

        os.makedirs(os.path.dirname(full_file_path), exist_ok=True)

        file.save(full_file_path)

        new_file = Docs2(
            filename=filename_to_save,
            file_path=relative_path_from_static,
            upload_by_id=current_user.id,
            ticket_id=ticket.id
        )

        try:
            db.session.add(new_file)
            db.session.commit()
            flash(
                f'File {filename_to_save} berhasil diupload ke Lender Docs (menggantikan file sebelumnya jika ada).', 'success')
        except Exception as e:
            db.session.rollback()
            if os.path.exists(full_file_path):
                os.remove(full_file_path)
            flash(f'Gagal mengupload file: {str(e)}', 'danger')
    else:
        flash('Tipe file tidak diizinkan.', 'danger')

    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


@app.route('/save-tenor/<int:ticket_id>', methods=['POST'])
@login_required
def save_tenor(ticket_id):
    """Simpan atau perbarui data tenor untuk ticket tertentu"""
    ticket = Ticket.query.get_or_404(ticket_id)

    selected_order_number = request.form.get('order_number', '').strip()
    nominal_str = request.form.get('nominal', '').strip()
    tenor_str = request.form.get('tenor', '').strip()
    tanggal_ovd_str = request.form.get('tanggalPengaduan', '').strip()

    if not (selected_order_number and nominal_str and tenor_str and tanggal_ovd_str):
        flash(
            'Order Number, Nominal, Tenor, dan Tanggal Jatuh Tempo wajib diisi.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    order_numbers_list = [num.strip() for num in ticket.order_number.split(
        ',')] if ticket.order_number else []
    if selected_order_number not in order_numbers_list:
        flash('Order Number yang dipilih tidak valid untuk ticket ini.', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    try:
        nominal = float(nominal_str)
        tenor = int(tenor_str)
        if tenor < 1 or tenor > 12:
            raise ValueError("Tenor harus antara 1 dan 12.")
        tanggal_ovd = datetime.strptime(tanggal_ovd_str, '%Y-%m-%d').date()
    except ValueError as e:
        flash(f'Format input tidak valid: {str(e)}', 'danger')
        return redirect(url_for('ticket_detail', ticket_id=ticket_id))

    nominal_per_tenor = nominal / tenor

    existing_tenor = Tenor.query.filter_by(
        ticket_id=ticket.id, nomor_kontrak=selected_order_number).first()

    if existing_tenor:
        db.session.delete(existing_tenor)

    new_tenor = Tenor(
        nomor_kontrak=selected_order_number,
        total_nominal=nominal,
        total_nominal_akhir=nominal,
        ticket_id=ticket.id
    )

    for i in range(1, tenor + 1):
        setattr(new_tenor, f'tenor_{i}', i)
        setattr(new_tenor, f'nominal_tenor_{i}', nominal_per_tenor)
        ovd_date = add_months(tanggal_ovd, i - 1)
        setattr(new_tenor, f'ovd_{i}', ovd_date)

    try:
        db.session.add(new_tenor)
        db.session.commit()
        flash('Data tenor berhasil disimpan.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menyimpan data tenor: {str(e)}', 'danger')

    return redirect(url_for('ticket_detail', ticket_id=ticket_id))


def add_months(source_date, months):
    """Menambahkan bulan ke tanggal, menyesuaikan tanggal jika melewati akhir bulan."""
    month = source_date.month - 1 + months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, calendar.monthrange(year, month)[1])
    return source_date.replace(year=year, month=month, day=day)


@app.route('/view-kontrak/<int:id>', methods=['GET'])
@login_required
def view_kontrak(id):
    tenor = Tenor.query.get_or_404(id)

    def build_row(t):
        def fmt_nominal(v):
            if v is None:
                return None
            try:
                return '{:,.0f}'.format(float(v)).replace(',', '.')
            except Exception:
                return None

        row = {
            'nomor_kontrak': (t.nomor_kontrak or '').strip(),
            'total_nominal_pengembalian': float(t.total_nominal) if t.total_nominal is not None else None,
            'total_nominal_akhir': fmt_nominal(t.total_nominal_akhir),
            'cicilan': []
        }

        jumlah_tenor_aktif = 0
        for i in range(1, 13):
            if (getattr(t, f'tenor_{i}', None) is not None and
                getattr(t, f'nominal_tenor_{i}', None) is not None and
                    getattr(t, f'ovd_{i}', None) is not None):
                jumlah_tenor_aktif = i
            else:
                break
        row['jumlah_tenor_aktif'] = jumlah_tenor_aktif

        for i in range(1, 13):
            tenor_i = getattr(t, f'tenor_{i}', None)
            nominal_i = getattr(t, f'nominal_tenor_{i}', None)
            ovd_i = getattr(t, f'ovd_{i}', None)

            if tenor_i is not None and nominal_i is not None and ovd_i is not None:
                row['cicilan'].append({
                    'number': i,
                    'tanggal_jatuh_tempo': ovd_i.strftime('%d/%m/%Y'),
                    'nominal': float(nominal_i)
                })
            else:
                row['cicilan'].append(None)

        return row

    row_data = build_row(tenor)
    return render_template('view_tenor.html', user=current_user, tenor_data=[row_data])


def format_rupiah_id(n):
    if n is None:
        return ''
    try:
        if isinstance(n, Decimal):
            n = int(n)
        else:
            n = int(float(n))
        s = f"{n:,}".replace(",", ".")
        return f"Rp{s}"
    except Exception:
        return ''


def fmt_tgl(d):
    return d.strftime("%d/%m/%Y") if d else ''


@app.route('/view-all-kontrak/<int:ticket_id>', methods=['GET'])
@login_required
def view_all_kontrak(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    tenors = Tenor.query.filter_by(ticket_id=ticket_id).all()

    if request.args.get('export') == 'xlsx':
        only_order = (request.args.get('order') or '').strip()

        def is_nullish(v):
            return (v is None) or (isinstance(v, str) and v.strip().upper() in {"", "NULL", "(NULL)"})

        ORANGE = "DEB887"
        BROWN = "8B4513"
        fill_head = PatternFill("solid", fgColor=ORANGE)
        font_head = Font(bold=True)
        align_center = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center")
        side_thin = Side(border_style="thin", color=BROWN)
        border_all = Border(left=side_thin, right=side_thin,
                            top=side_thin, bottom=side_thin)
        nf_money = u'\"Rp\"#,##0'
        nf_date = "DD/MM/YYYY"

        wb = Workbook()
        ws = wb.active
        ws.title = "Tenor"

        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['H'].width = 22

        row_cursor = 1

        for t in tenors:
            nomor = (t.nomor_kontrak or '').strip()
            if only_order and nomor != only_order:
                continue

            items = []
            for i in range(1, 13):
                tenor_i = getattr(t, f"tenor_{i}", None)
                if is_nullish(tenor_i):
                    continue
                ovd_i = getattr(t, f"ovd_{i}", None)
                nom_i = getattr(t, f"nominal_tenor_{i}", None)
                items.append({
                    "idx": i,
                    "tgl": ovd_i if isinstance(ovd_i, (date, datetime)) else None,
                    "nom": float(nom_i) if nom_i is not None else None
                })

            n = len(items)
            groups = ceil(max(1, n) / 3)
            total_rows = groups * 3

            ws.merge_cells(start_row=row_cursor, start_column=1,
                           end_row=row_cursor + total_rows - 1, end_column=1)
            a = ws.cell(row=row_cursor, column=1,
                        value=f"Nomor Kontrak\n\n{nomor}")
            a.fill = fill_head
            a.font = font_head
            a.alignment = align_center
            a.border = border_all

            for g in range(groups):
                base_c = 2
                base_r = row_cursor + g * 3

                for k in range(3):
                    cicil_no = (g * 3) + k + 1
                    ws.merge_cells(start_row=base_r, start_column=base_c + (k * 2),
                                   end_row=base_r, end_column=base_c + (k * 2) + 1)
                    ch = ws.cell(row=base_r, column=base_c +
                                 (k * 2), value=f"Cicilan ke {cicil_no}")
                    ch.fill = fill_head
                    ch.font = font_head
                    ch.alignment = align_center
                    ch.border = border_all

                for k in range(3):
                    h1 = ws.cell(row=base_r + 1, column=base_c +
                                 (k * 2), value="Tanggal Jatuh Tempo")
                    h2 = ws.cell(row=base_r + 1, column=base_c +
                                 (k * 2) + 1, value="Nominal")
                    for cc in (h1, h2):
                        cc.fill = fill_head
                        cc.font = font_head
                        cc.alignment = align_center
                        cc.border = border_all

                for k in range(3):
                    idx = g * 3 + k
                    c_tgl = ws.cell(row=base_r + 2, column=base_c + (k * 2))
                    c_nom = ws.cell(
                        row=base_r + 2, column=base_c + (k * 2) + 1)
                    if idx < n and items[idx]["tgl"]:
                        c_tgl.value = items[idx]["tgl"]
                        c_tgl.number_format = nf_date
                    else:
                        c_tgl.value = "-"
                    if idx < n and items[idx]["nom"] is not None:
                        c_nom.value = items[idx]["nom"]
                        c_nom.number_format = nf_money
                        c_nom.alignment = align_right
                    else:
                        c_nom.value = "-"
                        c_nom.alignment = align_center
                    c_tgl.alignment = align_center
                    c_tgl.border = border_all
                    c_nom.border = border_all

                for rr in range(base_r, base_r + 3):
                    gcell = ws.cell(row=rr, column=7)
                    gcell.border = border_all

            label_top = row_cursor
            label_bottom = row_cursor + total_rows - 2
            value_row = row_cursor + total_rows - 1
            ws.merge_cells(start_row=label_top, start_column=8,
                           end_row=label_bottom, end_column=8)
            lab = ws.cell(row=label_top, column=8,
                          value="Total nominal pengembalian")
            lab.fill = fill_head
            lab.font = font_head
            lab.alignment = align_center
            lab.border = border_all

            total_akhir = float(
                t.total_nominal_akhir) if t.total_nominal_akhir is not None else None
            val = ws.cell(row=value_row, column=8)
            if total_akhir is not None:
                val.value = total_akhir
                val.number_format = nf_money
            else:
                val.value = "-"
            val.fill = fill_head
            val.font = font_head
            val.alignment = align_center
            val.border = border_all

            row_cursor += total_rows

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"tenor_ticket_{ticket_id}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    rows = []
    max_tenor_aktif = 0

    for t in tenors:
        aktif = 0
        for i in range(1, 13):
            tenor_i = getattr(t, f'tenor_{i}', None)
            nominal_i = getattr(t, f'nominal_tenor_{i}', None)
            ovd_i = getattr(t, f'ovd_{i}', None)
            if tenor_i is not None and nominal_i is not None and ovd_i is not None:
                aktif = i
            else:
                break

        max_tenor_aktif = max(max_tenor_aktif, aktif)
        cicilan = []
        for i in range(1, 13):
            ovd_i = getattr(t, f"ovd_{i}", None)
            nom_i = getattr(t, f"nominal_tenor_{i}", None)
            if ovd_i and nom_i:
                tgl_str = fmt_tgl(ovd_i)
                nom_str = format_rupiah_id(float(nom_i))
            else:
                tgl_str = "-"
                nom_str = "-"
            cicilan.append({"tgl": tgl_str, "nom": nom_str})

        total_display = format_rupiah_id(
            float(t.total_nominal)) if t.total_nominal else "-"
        total_akhir_display = format_rupiah_id(
            float(t.total_nominal_akhir)) if t.total_nominal_akhir else "-"

        rows.append({
            "nomor_kontrak": (t.nomor_kontrak or "").strip(),
            "total": total_display,
            "total_nominal_akhir": total_akhir_display,
            "c": cicilan,
        })

    return render_template(
        'view_all_tenor.html',
        user=current_user,
        ticket=ticket,
        rows=rows,
        max_tenor_aktif=max_tenor_aktif
    )


def sla_class(sla: int | None) -> str:
    if sla is None:
        return "!text-default-400"
    if sla >= 8:
        return "!text-success"
    if sla >= 4:
        return "!text-info"
    return "!text-danger"


def ovd_class(_):
    return "!text-danger text-center"


@app.route('/api/calendar-events', methods=['GET'])
@login_required
def api_calendar_events():
    tickets = Ticket.query.all()
    today_jkt = get_jakarta_time().date()

    events = []
    for t in tickets:

        if t.status == 3:
            continue
        
        if t.case_progress != 1:
            continue

        sla_days = t.sla or 0

        if sla_days <= 0:
            base_date = (
                t.tanggal_pengerjaan or t.tanggal_pengaduan or get_jakarta_time()).date()
            due_date = base_date + timedelta(days=10)
        else:
            due_date = today_jkt + timedelta(days=sla_days)

        events.append({
            "id": t.id,
            "title": t.no_ticket,
            "start": due_date.isoformat(),
            "allDay": True,
            "className": sla_class(sla_days),
        })

    return jsonify(events)


@app.route('/api/calendar-ovd', methods=['GET'])
@login_required
def api_calendar_ovd():
    tenors = Tenor.query.all()
    events = []

    for tenor in tenors:

        if tenor.ticket and tenor.ticket.status == 3:
            continue

        if tenor.ticket and tenor.ticket.case_progress != 2:
            continue

        for i in range(1, 13):
            ovd_value = getattr(tenor, f"ovd_{i}")
            if ovd_value:
                events.append({
                    "id": tenor.ticket_id,
                    "title": f"{tenor.nomor_kontrak}",
                    "start": ovd_value.strftime("%Y-%m-%d"),
                    "nomor_kontrak": tenor.nomor_kontrak,
                    "tenor": f"Tenor ke-{i}",
                    "className": ovd_class(ovd_value)
                })

    return jsonify(events)

@app.route('/calendar', methods=['GET'])
@login_required
def calendar_view():
    return render_template('calendar.html', user=current_user)


@app.route('/calendar-ovd', methods=['GET'])
@login_required
def calendar_ovd():
    return render_template('calendar_ovd.html', user=current_user)


@app.route('/close/<int:id>', methods=['POST'])
@login_required
def close(id):
    ticket = Ticket.query.get_or_404(id)

    ticket.status = 3
    ticket.close_date = datetime.now()
    ticket.close_by = current_user.id

    db.session.commit()

    flash(
        f"Case #{ticket.no_ticket} telah ditutup oleh {current_user.name}.", "success")
    return redirect(url_for('ticket_detail', ticket_id=id))


@app.route('/move-case/<int:id>', methods=['POST'])
@login_required
def move_case(id):

    ticket = Ticket.query.get_or_404(id)

    ticket.case_progress = 2
    ticket.move_date = datetime.now()
    ticket.move_by = current_user.id

    db.session.commit()

    flash(f"Case #{ticket.no_ticket} move by {current_user.name}.", "success")
    return redirect(url_for('ticket_detail', ticket_id=id))


@app.route('/chat')
@login_required
def chat():
    users = User.query.filter(User.id != current_user.id).all()
    return render_template('chat.html', user=current_user, users=users)

@app.route('/collection-template', methods=['GET', 'POST'])
@login_required
def collection_template():

    if request.method == 'POST':
        title = request.form.get('title')
        detail = request.form.get('detail')

        if not title or not detail:
            flash('Title dan Detail wajib diisi', 'danger')
            return redirect(url_for('collection_template'))

        template = CollectionTemplate(
            title=title,
            detail=detail,
            status=1
        )

        db.session.add(template)
        db.session.commit()

        flash('Template berhasil ditambahkan', 'success')
        return redirect(url_for('collection_template'))

    templates = CollectionTemplate.query.order_by(CollectionTemplate.id.asc()).all()

    return render_template(
        'template_penagihan.html',
        user=current_user,
        templates=templates
    )

@app.route('/collection-template/delete', methods=['POST'])
@login_required
def delete_collection_template():
    template = CollectionTemplate.query.get_or_404(request.form.get('id'))
    db.session.delete(template)
    db.session.commit()

    flash('Template berhasil dihapus', 'success')
    return redirect(url_for('collection_template'))

@app.route('/reply-template', methods=['GET', 'POST'])
@login_required
def reply_template():

    if request.method == 'POST':
        title = request.form.get('title')
        detail = request.form.get('detail')

        if not title or not detail:
            flash('Title dan Detail wajib diisi', 'danger')
            return redirect(url_for('reply_template'))

        template = ReplyTemplate(
            title=title,
            detail=detail,
            status=1
        )

        db.session.add(template)
        db.session.commit()

        flash('Template berhasil ditambahkan', 'success')
        return redirect(url_for('reply_template'))

    templates = ReplyTemplate.query.order_by(ReplyTemplate.id.asc()).all()

    return render_template('template_balasan.html', user=current_user, templates=templates)

@app.route('/reply-template/delete', methods=['POST'])
@login_required
def delete_reply_template():
    template = ReplyTemplate.query.get_or_404(request.form.get('id'))
    db.session.delete(template)
    db.session.commit()

    flash('Template berhasil dihapus', 'success')
    return redirect(url_for('reply_template'))

# @app.route('/chat/<int:user_id>')
# @login_required
# def chat_with(user_id):
#     other_user = User.query.get_or_404(user_id)
#     messages = Message.query.filter(
#         ((Message.sender_id == current_user.id) & (Message.receiver_id == user_id)) |
#         ((Message.sender_id == user_id) & (Message.receiver_id == current_user.id))
#     ).order_by(Message.timestamp.asc()).all()

#     return jsonify({
#         'other_user': {
#             'id': other_user.id,
#             'name': other_user.name,
#             'email': other_user.email
#         },
#         'messages': [
#             {
#                 'sender_id': m.sender_id,
#                 'receiver_id': m.receiver_id,
#                 'message': m.message,
#                 'timestamp': m.timestamp.strftime('%H:%M')
#             } for m in messages
#         ]
#     })


# @app.route('/send_message', methods=['POST'])
# @login_required
# def send_message():
#     receiver_id = request.form.get('receiver_id')
#     message_text = request.form.get('message')

#     if not receiver_id or not message_text:
#         return jsonify({'error': 'Missing fields'}), 400

#     message = Message(sender_id=current_user.id,
#                       receiver_id=receiver_id, message=message_text)
#     db.session.add(message)
#     db.session.commit()

#     return jsonify({'success': True, 'timestamp': message.timestamp.strftime('%H:%M')})


@app.route('/download-template')
def download_template():

    return send_from_directory(directory='static/files', path='template.xlsx', as_attachment=True)


@app.route('/import-tickets', methods=['POST'])
def import_tickets():
    if 'file' not in request.files:
        flash('Tidak ada file yang diupload', 'error')
        return redirect(request.referrer)

    file = request.files['file']
    product_id = request.form.get('product_id')

    if not product_id or not product_id.isdigit():
        flash('Pilih produk terlebih dahulu', 'error')
        return redirect(request.referrer)

    product_id = int(product_id)

    if file.filename == '':
        flash('File tidak valid', 'error')
        return redirect(request.referrer)

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Format file harus Excel (.xlsx/.xls)', 'error')
        return redirect(request.referrer)

    try:
        df = pd.read_excel(file)

        required_cols = [
            'tanggal_pengaduan', 'nama', 'nomor_pengajuan', 'nomor_aktif',
            'email', 'order_number', 'kanal_pengaduan',
            'detail_pengaduan', 'detail_problem'
        ]

        if not all(col in df.columns for col in required_cols):
            flash('File Excel tidak memiliki kolom yang diperlukan', 'error')
            return redirect(request.referrer)

        jakarta_tz = pytz.timezone('Asia/Jakarta')
        now_jakarta = datetime.now(jakarta_tz)

        product = Product.query.get(product_id)
        if not product:
            flash('Produk tidak ditemukan', 'error')
            return redirect(request.referrer)

        for _, row in df.iterrows():
            ticket = Ticket(
                no_ticket=generate_ticket_number(),
                tanggal_pengaduan=now_jakarta,
                tanggal_pengerjaan=now_jakarta,
                nama=row['nama'],
                phone_pengajuan=str(row['nomor_pengajuan']).strip(),
                phone_aktif=str(row['nomor_aktif']).strip(),
                email=row['email'],
                order_number=str(row['order_number']) if pd.notna(
                    row['order_number']) else None,
                pic_handle_id=current_user.id,
                product_id=product_id,
                kanal_pengaduan=row['kanal_pengaduan'],
                detail_pengaduan=row['detail_pengaduan'],
                detail_problem=row['detail_problem'],
                tipe_pengaduan=product.complaint_type,
                status=1,
                case_progress=1
            )
            db.session.add(ticket)

        db.session.commit()
        flash(f'Berhasil mengimpor {len(df)} tiket', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error saat import: {str(e)}', 'error')
        print("Error:", e)

    return redirect(request.referrer)


@app.before_request
def create_tables_if_not_exists():
    if not hasattr(g, 'tables_created'):
        db.create_all()
        g.tables_created = True

@app.route('/export-ticket', methods=['POST'])
def export_ticket():
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    if not start_date or not end_date:
        flash("Start Date dan End Date wajib diisi.", "danger")
        return redirect(url_for('dashboard'))

    try:
        start_date_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_dt = datetime.strptime(end_date, "%Y-%m-%d")

        end_date_dt = end_date_dt.replace(hour=23, minute=59, second=59)
    except Exception as e:
        flash("Format tanggal tidak valid.", "danger")
        return redirect(url_for('dashboard'))

    tickets = Ticket.query.filter(
        Ticket.tanggal_pengaduan >= start_date_dt,
        Ticket.tanggal_pengaduan <= end_date_dt
    ).all()

    if not tickets:
        flash("Tidak ada data pada range tersebut.", "warning")
        return redirect(url_for('dashboard'))

    data = []

    status_map = {
        1: "Open",
        2: "On Progress",
        3: "Close"
    }

    for t in tickets:
        data.append({
            "No Ticket": t.no_ticket,
            "Tanggal Tiket Masuk": t.tanggal_pengaduan,
            "Nama Nasabah": t.nama,
            "Nomor Nasabah": t.phone_pengajuan,
            "Email": t.email,
            "Order Number": t.order_number,
            "PIC": t.pic_handle.name if t.pic_handle else "-",
            "Detail Problem": t.detail_problem,
            "Tipe Pengaduan": t.product.complaint_type if t.product else "",
            "Kanal Pengaduan": t.kanal_pengaduan,
            "Detail Pengaduan": t.detail_pengaduan,
            "Status": status_map.get(t.status, "Unknown"),
        })


    df = pd.DataFrame(data)

    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    filename = f"export_ticket_{start_date}_to_{end_date}.xlsx"

    return send_file(output,
                     download_name=filename,
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# =========================
# CONFIG TEMPLATE PATH
# =========================

import io
from decimal import Decimal, InvalidOperation
from datetime import datetime, time
from openpyxl import load_workbook

EXCEL_TEMPLATE_PATH = os.path.join(app.root_path, "static", "files", "excel.xlsx")

MONTH_SHEETS_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]

RUPIAH_FORMAT = '"Rp"#,##0;[Red]-"Rp"#,##0'
SELISIH_FONT = Font(color="FF0000", italic=True)

def _split_csv_like(value: str) -> list[str]:
    """Split string 'a,b,c' -> ['a','b','c'] trimming, ignore kosong."""
    if not value:
        return []
    parts = [p.strip() for p in str(value).split(",")]
    return [p for p in parts if p]


def _parse_decimal_maybe(value: str) -> Decimal:
    """
    Parse angka dari string.
    Mendukung: '500000', '37,500.00', '37.500,00' (heuristik sederhana).
    """
    if value is None:
        return Decimal("0")

    s = str(value).strip()
    if not s:
        return Decimal("0")

    s = re.sub(r"[^\d,.\-]", "", s)

    if "." in s and "," in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if "," in s and "." not in s:
            chunks = s.split(",")
            if len(chunks[-1]) == 3:
                s = s.replace(",", "")
            else:
                s = s.replace(",", ".")

    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _count_order_numbers(ticket: Ticket) -> int:
    return len(_split_csv_like(ticket.order_number or ""))


def _sum_nominal_order(ticket: Ticket) -> Decimal:
    vals = _split_csv_like(ticket.nominal_order or "")
    total = Decimal("0")
    for v in vals:
        total += _parse_decimal_maybe(v)
    return total


def _clear_sheet_from_row(ws, start_row: int = 4):
    """Kosongkan isi sheet mulai start_row sampai bawah."""
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < start_row:
        return
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def _parse_date_param(param_name: str):
    """
    Ambil date dari query string 'YYYY-MM-DD'.
    Return datetime atau None.
    """
    v = request.args.get(param_name)
    if not v:
        return None
    return datetime.strptime(v, "%Y-%m-%d")

@app.route("/export/restruktur-collection", methods=["GET"])
def export_restruktur_collection():
    """
    Export excel berdasarkan template excel.xlsx.
    Query params:
      - start_date=YYYY-MM-DD (optional)
      - end_date=YYYY-MM-DD (optional)
    """
    start_dt = _parse_date_param("start_date")
    end_dt = _parse_date_param("end_date")

    q = db.session.query(Ticket).filter(Ticket.case_progress == 2)

    if start_dt:
        q = q.filter(Ticket.tanggal_pengaduan >= start_dt)

    if end_dt:
        end_of_day = datetime.combine(end_dt.date(), time.max)
        q = q.filter(Ticket.tanggal_pengaduan <= end_of_day)

    tickets = q.all()

    ticket_ids = [t.id for t in tickets]
    tenors_by_ticket = {}
    if ticket_ids:
        tenors = db.session.query(Tenor).filter(Tenor.ticket_id.in_(ticket_ids)).all()
        for tr in tenors:
            tenors_by_ticket.setdefault(tr.ticket_id, []).append(tr)

    metrics = {}

    for t in tickets:
        if not t.tanggal_pengaduan:
            continue

        dt = t.tanggal_pengaduan.date()
        month = dt.month
        key = (month, dt)

        if key not in metrics:
            metrics[key] = {
                "fu_total_user": 0,
                "fu_total_data": 0,
                "fu_total_amount": Decimal("0"),
                "rp_total_user": 0,
                "rp_total_data": 0,
                "rp_total_amount": Decimal("0"),
            }

        metrics[key]["fu_total_user"] += 1
        metrics[key]["fu_total_data"] += _count_order_numbers(t)
        metrics[key]["fu_total_amount"] += _sum_nominal_order(t)

        t_tenors = tenors_by_ticket.get(t.id, [])
        had_repayment = False
        repaid_tenor_count = 0
        repaid_amount = Decimal("0")

        for tr in t_tenors:
            total_nominal = Decimal(tr.total_nominal or 0)
            total_akhir = Decimal(tr.total_nominal_akhir or 0)
            if total_akhir < total_nominal:
                had_repayment = True
                repaid_tenor_count += 1
                repaid_amount += (total_nominal - total_akhir)

        if had_repayment:
            metrics[key]["rp_total_user"] += 1

        metrics[key]["rp_total_data"] += repaid_tenor_count
        metrics[key]["rp_total_amount"] += repaid_amount

    wb = load_workbook(EXCEL_TEMPLATE_PATH)

    base_ws = wb.active
    base_title = base_ws.title

    month_ws = {}
    for i, month_name in enumerate(MONTH_SHEETS_ID, start=1):
        if month_name in wb.sheetnames:
            ws = wb[month_name]
        else:
            ws = wb.copy_worksheet(base_ws)
            ws.title = month_name

        _clear_sheet_from_row(ws, start_row=4)
        month_ws[i] = ws

    if base_title not in MONTH_SHEETS_ID and base_title in wb.sheetnames:
        pass

    for (m, dt), v in sorted(metrics.items(), key=lambda x: (x[0][0], x[0][1])):
        ws = month_ws[m]

        row = 4
        while ws.cell(row=row, column=1).value not in (None, ""):
            row += 1

        fu_amt = v["fu_total_amount"]
        rp_amt = v["rp_total_amount"]
        selisih = fu_amt - rp_amt

        ws.cell(row=row, column=1).value = dt.strftime("%Y-%m-%d")
        ws.cell(row=row, column=2).value = v["fu_total_user"]
        ws.cell(row=row, column=3).value = v["fu_total_data"]
        cell_total_amount = ws.cell(row=row, column=4)
        cell_total_amount.value = float(fu_amt)
        cell_total_amount.number_format = RUPIAH_FORMAT

        ws.cell(row=row, column=5).value = v["rp_total_user"]
        ws.cell(row=row, column=6).value = v["rp_total_data"]
        cell_repayment = ws.cell(row=row, column=7)
        cell_repayment.value = float(rp_amt)
        cell_repayment.number_format = RUPIAH_FORMAT

        cell_selisih = ws.cell(row=row, column=8)
        cell_selisih.value = float(selisih)
        cell_selisih.number_format = RUPIAH_FORMAT
        cell_selisih.font = SELISIH_FONT

    if "Sheet1" in wb.sheetnames:
        wb.remove(wb["Sheet1"])

    for sheet_name in wb.sheetnames[:]:
        if sheet_name not in MONTH_SHEETS_ID:
            if len(wb.sheetnames) > 1:
                wb.remove(wb[sheet_name])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    if start_dt or end_dt:
        sd = start_dt.strftime("%Y%m%d") if start_dt else "ALL"
        ed = end_dt.strftime("%Y%m%d") if end_dt else "ALL"
        filename = f"RESTRUKTUR_COLLECTION_{sd}_{ed}.xlsx"
    else:
        filename = f"RESTRUKTUR_COLLECTION_{suffix}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

log = logging.getLogger('werkzeug')
log.setLevel(logging.WARNING)

if __name__ == '__main__':
    scheduler = BackgroundScheduler(timezone='Asia/Jakarta')
    scheduler.add_job(
        func=reduce_sla_daily,
        trigger="cron",
        hour=0,
        minute=0,
        id='reduce_sla_job'
    )
    scheduler.start()
    print("Program Running⏳")

    try:
        app.run(debug=True, port=5006, host='0.0.0.0')
    finally:
        atexit.register(lambda: scheduler.shutdown())
